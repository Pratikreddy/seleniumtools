from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from datetime import date
import time
from selenium.webdriver.chrome.options import Options
import os
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
#import fitz
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import date, timedelta
from docx import Document
from fpdf import FPDF

s = time.time()
today = date.today()
options = Options()
# options.add_argument('--headless')
driver = None



def browser_config(default_path, chrome_path, headless=None):
    global driver
    options = webdriver.ChromeOptions()
    options.add_experimental_option("prefs", {
        "download.default_directory": default_path,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True
    })
    
    if headless == "Yes":
        options.add_argument("--headless=new") 
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--remote-debugging-port=9222")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--disable-extensions")
        options.add_argument("--no-start-maximized")
    
    driver = webdriver.Chrome(executable_path=chrome_path, options=options)



def login(url,userID,password,id_element,pass_element):
    try:
        driver.get(url)
        driver.maximize_window()
        WebDriverWait(driver,60).until(EC.presence_of_element_located((By.XPATH,id_element)))
        user_input=driver.find_element(By.XPATH,(id_element))
        user_input.send_keys(userID)
        pass_input=driver.find_element(By.XPATH,(pass_element))
        pass_input.send_keys(password)
        pass_input.send_keys(Keys.RETURN)
        # WebDriverWait(driver,60).until(EC.invisibility_of_element(By.XPATH,pass_input))
    except:
        print("login failed")
        
def open_url(url,path):
    try:
        driver.get(url)
        driver.maximize_window()
        WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH,path)))
        driver.find_element(By.XPATH,(path)).click()
        print("clicked on element ",path)
    except:
        print(" click on element failed for  ",path)
       
def open_webpage(url):    
    driver.get(url)
    driver.maximize_window()
    time.sleep(10)
    
    
def click(path):
    try:
        WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH,path)))
        driver.find_element(By.XPATH,(path)).click()
        print("clicked on element ",path)
    except:
        print(" click on element failed for  ",path)

def handle_files(default_path,report_name):
    os.chdir(default_path)
    all_files_nonsorted = os.listdir(os.getcwd())    
    for i in os.listdir(default_path):
     ext = i.split('.')
     while ext[-1] == 'crdownload':
        for i in os.listdir(default_path):
            print("download in progress")
            time.sleep(2)
            ext = i.split('.')                                
    all_files = sorted(os.listdir(os.getcwd()),key=os.path.getatime)
    downloaded_file = all_files[-1]
    try:
        os.rename(downloaded_file, (str(today)+"_"+report_name))
    except:
        path_file = os.path.join(default_path,(str(today)+"_"+report_name))
        os.remove(path_file)
        os.rename(downloaded_file, (str(today)+"_"+report_name))

def clear_input_field(element_xpath):
    element = driver.find_element(By.XPATH, element_xpath)
    element.clear()

def type_into(element_path,text):   
    WebDriverWait(driver,60).until(EC.presence_of_element_located((By.XPATH,element_path)))
    user_input=driver.find_element(By.XPATH,(element_path))
    user_input.send_keys(text)

def handle_popup():
    original_window = driver.current_window_handle
    popup_window = None
    WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(2))
    for window_handle in driver.window_handles:
        if window_handle != original_window:
            popup_window = window_handle
            break
    driver.switch_to.window(popup_window)
    driver.close()
    driver.switch_to.window(original_window)

def select_from_dropdown(dropdown_path, option_text):
    dropdown = Select(driver.find_element(By.XPATH,(dropdown_path)))
    dropdown.select_by_visible_text(option_text)

def is_element_visible(element_xpath):
    try:
        element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, element_xpath)))
        return True
    except:
        return False

def scroll_to_element(element_xpath):
    element = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, element_xpath)))
    driver.execute_script("arguments[0].scrollIntoView();", element)

def get_current_url():
    return driver.current_url

def wait_for_element_to_present(element_xpath):
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, element_xpath)))


def wait_for_element_to_disappear(element_xpath):
    try:
        WebDriverWait(driver, 60).until_not(EC.presence_of_element_located((By.XPATH, element_xpath)))
    except:
        print("Element {} did not disappear".format(element_xpath))

def scroll_to_top():
    driver.execute_script("window.scrollTo(0, 0);")

def scroll_to_bottom():
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

def get_element_text(element_xpath):
    element = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, element_xpath)))
    return element.text

def get_element_attribute(element_xpath, attribute_name):
    element = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, element_xpath)))
    return element.get_attribute(attribute_name)

#new
def find_elements(xpath, timeout=30):
    """
    Finds all the elements that match the given XPath expression.
    Returns a list of matching elements.
    """
    return WebDriverWait(driver, timeout).until(EC.presence_of_all_elements_located((By.XPATH, xpath)))


def wait(timeout):
    time.sleep(timeout)

def wait_for_page_title(page_title):
    WebDriverWait(driver, 60).until(EC.title_contains(page_title))

def take_screenshot(file_path):
    driver.save_screenshot(file_path)

def go_back():
    driver.back()

def switch_to_frame(frame_xpath):
    frame = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, frame_xpath)))
    driver.switch_to.frame(frame)

def switch_to_default_content():
    driver.switch_to.default_content()

def switch_to_window_by_title(title):
    main_window = driver.current_window_handle
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if driver.title == title:
            return handle
    driver.switch_to.window(main_window)
    return None

def switch_to_new_window():
    handles = driver.window_handles
    driver.switch_to.window(handles[-1])

def switch_to_previous_window():
    main_window = driver.current_window_handle
    all_windows = driver.window_handles
    if len(all_windows) > 1:
        previous_window = all_windows[all_windows.index(main_window) - 1]
        driver.switch_to.window(previous_window)

def send_email(sender_email, sender_password, recipient_email, subject, body):
    try:
        message = MIMEMultipart()
        message['From'] = sender_email
        message['To'] = recipient_email
        message['Subject'] = subject
        message.attach(MIMEText(body, 'plain'))
        smtp_server = smtplib.SMTP('smtp.gmail.com', 587)
        smtp_server.starttls()
        smtp_server.login(sender_email, sender_password)
        smtp_server.sendmail(sender_email, recipient_email, message.as_string())
        smtp_server.quit()
        print("Email sent successfully")
    except Exception as e:
        print("Error sending email:", e)

def read_excell_cell(excel_file_path, sheet_name, cell_address):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet =workbook[sheet_name]
    cell_value = sheet[cell_address].value
    return str(cell_value)

def todays_date(format):
    return date.today().strftime(format)

def this_year():
    return date.today().strftime('%Y')

def this_month():
    return date.today().strftime('%m')

def this_day_in_name():
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_num = date.today().weekday()
    return days[day_num]
def this_month_in_name():
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    def this_month():
        return date.today().strftime('%m')
    month_num = int(this_month())
    return months[month_num - 1]
#new
def yesterdays_date(format):
    today = date.today()
    yesterday = today - timedelta(days=1)
    formatted_date = yesterday.strftime(format)
    return formatted_date

#new
def the_weekend(format):
    today = date.today()
    weekend = today - timedelta(days=3)
    formatted_date = weekend.strftime(format)
    return formatted_date

def the_week(format):
    today = date.today()
    weekend = today - timedelta(days=8)
    formatted_date = weekend.strftime(format)
    return formatted_date
    

# def get_text_from_pdf_to_excel(file_location,excel_location):
#     pdf_file = fitz.open(file_location)
#     no_pages = pdf_file.page_count
#     wb = Workbook()
#     for i in range(no_pages):
#         wb.create_sheet(title='Page {}'.format(i+1), index=i)
#     for page_num in range(no_pages):
#         page = pdf_file.load_page(page_num)
#         text = page.get_text("text")
#         sheet = wb['Page {}'.format(page_num+1)]
#         sheet['A1'] = text
#         font = Font(name='Courier New', size=10)

#         rows = sheet.max_row
#         cols = sheet.max_column
#         for row in range(1, rows+1):
#             for col in range(1, cols+1):
#                 cell = sheet.cell(row=row, column=col)
#                 cell.font = font
#     wb.save(excel_location)
#     pdf_file.close()

def combine_images_to_pdf(folder_path,output_filename='combined_pdf.pdf'):
    path = folder_path
    image_files = [os.path.join(path, f) for f in os.listdir(path) if f.endswith(".jpg") or f.endswith(".png")]
    pdf = FPDF()
    pdf.add_page()
    for image in image_files:
        pdf.image(image)
    out_path=os.path.join(folder_path,"combined_pdf_folder")
    print(out_path)
    os.chdir(folder_path)
    os.mkdir(out_path)
    out_pdf_path=os.path.join(out_path,output_filename)
    pdf.output(out_pdf_path, "F")
    

def combine_images_to_word(folder_path, output_filename='combined_word.docx'):
    if not output_filename.lower().endswith('.docx'):
        output_filename += '.docx'

    path = folder_path
    image_files = [os.path.join(path, f) for f in os.listdir(path) if f.endswith(".jpg") or f.endswith(".png")]

    doc = Document()

    for image_path in image_files:
        doc.add_picture(image_path)
        doc.add_page_break()

    out_path = os.path.join(folder_path, "combined_word_folder")
    os.makedirs(out_path, exist_ok=True)

    out_word_path = os.path.join(out_path, output_filename)
    doc.save(out_word_path)

    print(f"Word document saved as {out_word_path}")
    
def quit_driver():
    global driver
    if driver is not None:
        driver.quit()
        print("Driver quit successfully")
    else:
        print("No active driver found")
        


